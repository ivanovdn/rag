"""
Convert evaluation XLSX (4 sheets) into JSON datasets.

Usage:
    python scripts/convert_eval_xlsx.py eval_dataset.xlsx

Output:
    eval/datasets/retrieval_test.json
    eval/datasets/e2e_test.json
    eval/datasets/escalation_test.json
    eval/datasets/chatbot_test_cases.json
"""

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)


def cell_val(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def parse_pipe_list(text: str):
    """If text contains | pipes -> list of strings. Otherwise -> string."""
    if not text:
        return ""
    if "|" in text:
        return [item.strip() for item in text.split("|") if item.strip()]
    return text


def read_rows(ws, start_row: int = 2) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        cells = [cell_val(c) for c in row]
        while len(cells) < 6:
            cells.append("")
        if not cells[1]:
            continue
        # Skip header rows read as data
        if cells[0].lower() == "id" or cells[1].lower() in ("question",):
            continue
        rows.append(cells)
    return rows


def convert_tier1(ws) -> dict:
    rows = read_rows(ws)
    test_cases = []
    for cells in rows:
        tc = {
            "id": cells[0] or f"RET-{len(test_cases)+1:03d}",
            "question": cells[1],
            "expected_doc_id": cells[2],
            "expected_section_contains": cells[3],
            "expected_clause": cells[4],
            "expected_text_contains": parse_pipe_list(cells[5]) if len(cells) > 5 else "",
        }
        test_cases.append(tc)
    return {
        "metadata": {
            "name": "Retrieval Quality Test Set",
            "version": "1.0",
            "description": "Tests whether the correct policy chunk is retrieved.",
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_cases": len(test_cases),
        },
        "test_cases": test_cases,
    }


def convert_qa_tier(ws, id_prefix: str = "E2E", name: str = "End-to-End Q&A Test Set", description: str = "Tests full pipeline: retrieval + LLM generation.") -> dict:
    """Convert a Q&A sheet with columns: id, question, expected_answer, expected_doc, expected_section, expected_clause."""
    rows = read_rows(ws)
    test_cases = []
    for cells in rows:
        expected_answer = parse_pipe_list(cells[2])
        citations = []
        if cells[3]:
            cit = {"doc_id": cells[3]}
            if cells[4]:
                cit["section"] = cells[4]
            if cells[5]:
                cit["clause"] = cells[5]
            citations.append(cit)
        tc = {
            "id": cells[0] or f"{id_prefix}-{len(test_cases)+1:03d}",
            "question": cells[1],
            "expected_answer": expected_answer,
            "expected_citations": citations,
        }
        test_cases.append(tc)
    return {
        "metadata": {
            "name": name,
            "version": "1.0",
            "description": description,
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_cases": len(test_cases),
        },
        "test_cases": test_cases,
    }


def convert_chatbot(ws) -> dict:
    """Convert chatbot sheet — reads columns by header name, auto-generates IDs.

    Expected headers (any order):
        Expected Document, Expected Section, Expected Clause,
        Policy Rule, User Goal, Question, Expected Answer
    """
    # Read header row to build column map
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [cell_val(c).lower() for c in header_row]

    col = {}
    header_aliases = {
        "question": "question",
        "expected answer": "expected_answer",
        "expected document": "expected_doc",
        "expected section": "expected_section",
        "expected clause": "expected_clause",
    }
    for idx, h in enumerate(headers):
        for alias, key in header_aliases.items():
            if alias in h:
                col[key] = idx
                break

    missing = [k for k in ["question", "expected_answer", "expected_doc"] if k not in col]
    if missing:
        raise ValueError(f"Chatbot sheet missing required headers: {missing}. Found: {headers}")

    test_cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [cell_val(c) for c in row]
        while len(cells) <= max(col.values()):
            cells.append("")

        question = cells[col["question"]]
        if not question:
            continue

        expected_answer = parse_pipe_list(cells[col["expected_answer"]])
        expected_doc = cells[col.get("expected_doc", -1)] if "expected_doc" in col else ""
        expected_section = cells[col.get("expected_section", -1)] if "expected_section" in col else ""
        expected_clause = cells[col.get("expected_clause", -1)] if "expected_clause" in col else ""

        citations = []
        if expected_doc:
            cit = {"doc_id": expected_doc}
            if expected_section:
                cit["section"] = expected_section
            if expected_clause:
                cit["clause"] = expected_clause
            citations.append(cit)

        tc = {
            "id": f"CB-{len(test_cases)+1:03d}",
            "question": question,
            "expected_answer": expected_answer,
            "expected_citations": citations,
        }
        test_cases.append(tc)

    return {
        "metadata": {
            "name": "Chatbot Q&A Test Set",
            "version": "1.0",
            "description": "Tests chatbot answers against expected policy responses.",
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_cases": len(test_cases),
        },
        "test_cases": test_cases,
    }


def convert_tier3(ws) -> dict:
    rows = read_rows(ws)
    test_cases = []
    for cells in rows:
        tc = {
            "id": cells[0] or f"ESC-{len(test_cases)+1:03d}",
            "question": cells[1],
            "reason": cells[2],
            "category": cells[3] or "policy-gap",
            "should_escalate": str(cells[4]).upper() == "TRUE" if cells[4] else True,
        }
        test_cases.append(tc)
    return {
        "metadata": {
            "name": "Escalation Test Set",
            "version": "1.0",
            "description": "Questions the bot should NOT answer.",
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_cases": len(test_cases),
        },
        "test_cases": test_cases,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/convert_eval_xlsx.py <filled_template.xlsx>")
        sys.exit(1)
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)
    output_dir = Path("eval/datasets")
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    data_sheets = [s for s in wb.sheetnames if "instruct" not in s.lower()]
    converters = [
        ("retrieval_test.json", lambda ws: convert_tier1(ws)),
        ("e2e_test.json", lambda ws: convert_qa_tier(ws, id_prefix="E2E", name="End-to-End Q&A Test Set", description="Tests full pipeline: retrieval + LLM generation.")),
        ("escalation_test.json", lambda ws: convert_tier3(ws)),
        ("chatbot_test_cases.json", lambda ws: convert_chatbot(ws)),
    ]
    for i, (output_name, converter) in enumerate(converters):
        if i >= len(data_sheets):
            print(f"WARNING: No sheet found for {output_name}, skipping")
            continue
        ws = wb[data_sheets[i]]
        data = converter(ws)
        count = data["metadata"]["total_cases"]
        output_path = output_dir / output_name
        output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
        print(f"  {output_name}: {count} test cases -> {output_path}")
    wb.close()
    # Print first entry from each file
    print()
    for fname in ["retrieval_test.json", "e2e_test.json", "escalation_test.json", "chatbot_test_cases.json"]:
        fpath = output_dir / fname
        if fpath.exists():
            data = json.loads(fpath.read_text())
            cases = data.get("test_cases", [])
            if cases:
                print(f"--- {fname} (first entry) ---")
                print(json.dumps(cases[0], indent=2, ensure_ascii=False))
                print()


if __name__ == "__main__":
    main()
