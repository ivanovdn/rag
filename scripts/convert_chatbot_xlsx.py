"""
Convert chatbot test cases XLSX into evaluation JSON.

Usage:
    python scripts/convert_chatbot_xlsx.py chatbot_cases.xlsx
    python scripts/convert_chatbot_xlsx.py chatbot_cases.xlsx --sheet "Sheet1"

Output:
    eval/datasets/chatbot_test_cases.json
"""

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)


def cell_val(cell) -> str:
    if cell is None:
        return ""
    val = str(cell).strip()
    val = val.replace("\xa0", " ")
    val = re.sub(r"\s+", " ", val)
    return val


def is_header_row(cells: list[str]) -> bool:
    combined = " ".join(c.lower() for c in cells if c)
    header_markers = [
        "policy section reference",
        "user goal",
        "example user question",
        "positive examples",
        "negative examples",
        "policy rule",
    ]
    return any(marker in combined for marker in header_markers)


def read_rows(ws) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [cell_val(c) for c in row]
        while len(cells) < 7:
            cells.append("")
        # Skip empty rows (no question in col E)
        if not cells[4]:
            continue
        # Skip headers
        if is_header_row(cells):
            continue
        rows.append(cells)
    return rows


def convert(ws) -> dict:
    rows = read_rows(ws)
    positive_cases = []
    negative_cases = []

    for i, cells in enumerate(rows):
        policy = cells[0]
        section_ref = cells[1]
        policy_rule = cells[2]
        user_goal = cells[3]
        question = cells[4]
        positive_answer = cells[5]
        negative_answer = cells[6]
        base_id = f"TC-{i+1:03d}"

        if positive_answer:
            positive_cases.append(
                {
                    "id": f"{base_id}-POS",
                    "question": question,
                    "expected_answer": positive_answer,
                    "policy": policy,
                    "policy_section": section_ref,
                    "policy_rule": policy_rule,
                    "user_goal": user_goal,
                    "type": "positive",
                }
            )
        if negative_answer:
            negative_cases.append(
                {
                    "id": f"{base_id}-NEG",
                    "question": question,
                    "incorrect_answer": negative_answer,
                    "expected_behavior": "Bot must NOT give an answer similar to the negative example.",
                    "policy": policy,
                    "policy_section": section_ref,
                    "policy_rule": policy_rule,
                    "user_goal": user_goal,
                    "type": "negative",
                }
            )

    return {
        "metadata": {
            "name": "Chatbot Test Cases (Positive & Negative)",
            "version": "1.0",
            "description": "Paired test cases: correct answer vs incorrect answer.",
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_positive": len(positive_cases),
            "total_negative": len(negative_cases),
            "total_cases": len(positive_cases) + len(negative_cases),
        },
        "positive_cases": positive_cases,
        "negative_cases": negative_cases,
    }


def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python scripts/convert_chatbot_xlsx.py <chatbot_cases.xlsx> [--sheet SheetName]"
        )
        sys.exit(1)
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)
    sheet_name = None
    if "--sheet" in sys.argv:
        idx = sys.argv.index("--sheet")
        if idx + 1 < len(sys.argv):
            sheet_name = sys.argv[idx + 1]
    output_dir = Path("eval/datasets")
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"Using sheet: '{wb.sheetnames[0]}'")
    data = convert(ws)
    wb.close()
    output_path = output_dir / "chatbot_test_cases.json"
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"\n  chatbot_test_cases.json:")
    print(f"    Positive cases: {data['metadata']['total_positive']}")
    print(f"    Negative cases: {data['metadata']['total_negative']}")
    print(f"    Saved to:       {output_path}")
    if data["positive_cases"]:
        print(f"\n--- Sample positive case ---")
        print(json.dumps(data["positive_cases"][0], indent=2, ensure_ascii=False))
    if data["negative_cases"]:
        print(f"\n--- Sample negative case ---")
        print(json.dumps(data["negative_cases"][0], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
